# -*- coding: utf-8 -*-
"""合并两份面试词汇 Excel：全部汇总去重 + 保留双方工作表。"""
from __future__ import annotations

import json
import shutil
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
OLD = ROOT / "面试常用单词_Java_Python_FastAPI_LangChain.xlsx"
NEW = Path(r"e:\Users\i\Desktop\AI与Java后端面试词汇分类表.xlsx")
if not NEW.exists():
    NEW = ROOT / "AI与Java后端面试词汇分类表.xlsx"
DESKTOP_OLD = Path(r"e:\Users\i\Desktop\单词\面试常用单词_Java_Python_FastAPI_LangChain.xlsx")
DESKTOP_MERGED = Path(r"e:\Users\i\Desktop\单词\面试词汇合并表_AI_Java.xlsx")
DOCS = ROOT / "docs"

HEADER = [
    "分类", "单词", "用法/解释", "中文谐音", "音标",
    "发音", "技术栈", "重要程度", "对比词", "一句话回答",
]
LEVEL_MAP = {"核心": "必问", "高频": "常问", "了解": "了解"}


def cells(row) -> list[str]:
    vals = ["" if c is None else str(c).strip() for c in (row or [])[:10]]
    while len(vals) < 10:
        vals.append("")
    return vals


def copy_sheet(src_ws, dst_wb, title: str):
    if title in dst_wb.sheetnames:
        dst_wb.remove(dst_wb[title])
    dst = dst_wb.create_sheet(title)
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format
    for col_letter, dim in src_ws.column_dimensions.items():
        dst.column_dimensions[col_letter].width = dim.width
    return dst


def load_rows_from_sheet(ws, start_row=2) -> dict[str, list[str]]:
    out = {}
    for raw in ws.iter_rows(min_row=start_row, values_only=True):
        row = cells(raw)
        word = row[1]
        if not word:
            continue
        row[7] = LEVEL_MAP.get(row[7], row[7] or "了解")
        out[word.lower()] = row
    return out


def main() -> None:
    wb_old = openpyxl.load_workbook(OLD)
    wb_new = openpyxl.load_workbook(NEW)

    # 1) 合并全部汇总：先旧后新（同词以新表为准）
    merged: dict[str, list[str]] = {}
    if "全部汇总" in wb_old.sheetnames:
        merged.update(load_rows_from_sheet(wb_old["全部汇总"]))
    if "全部词汇" in wb_new.sheetnames:
        merged.update(load_rows_from_sheet(wb_new["全部词汇"]))

    # 重建全部汇总
    if "全部汇总" in wb_old.sheetnames:
        wb_old.remove(wb_old["全部汇总"])
    ws_all = wb_old.create_sheet("全部汇总", 0)
    ws_all.append(HEADER)
    for key in sorted(merged.keys()):
        ws_all.append(merged[key])

    # 2) 并入新表工作表（不覆盖社区添加单词）
    for name in wb_new.sheetnames:
        target = name
        if name == "全部词汇":
            target = "AI分类-全部词汇"
        if name == "分类目录":
            target = "AI分类-目录"
        if target == "社区添加单词":
            continue
        copy_sheet(wb_new[name], wb_old, target)

    # 3) 说明页追加合并说明
    note = "本文件由「面试常用单词…」与「AI与Java后端面试词汇分类表」合并。"
    if "说明" in wb_old.sheetnames:
        wb_old["说明"].append([note])
    else:
        ws = wb_old.create_sheet("说明")
        ws.append([note])

    out_path = OLD
    wb_old.save(out_path)
    wb_old.close()
    wb_new.close()

    # 导出站点数据
    wb = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
    rows_raw = list(wb["全部汇总"].iter_rows(values_only=True))
    wb.close()
    rows = []
    for r in rows_raw[1:]:
        if not r or r[1] is None or str(r[1]).strip() == "":
            continue
        rows.append([("" if (i >= len(r) or r[i] is None) else str(r[i]).strip()) for i in range(10)])
    DOCS.mkdir(exist_ok=True)
    (DOCS / "vocab-data.json").write_text(
        json.dumps({"header": HEADER, "rows": rows}, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    shutil.copy2(out_path, DOCS / "interview-vocab.xlsx")

    for dest in (DESKTOP_OLD, DESKTOP_MERGED):
        try:
            shutil.copy2(out_path, dest)
            print("copied", dest)
        except PermissionError:
            print("locked, skip", dest)

    # 同步一份合并表到仓库根目录别名
    alias = ROOT / "面试词汇合并表_AI_Java.xlsx"
    shutil.copy2(out_path, alias)
    print("merged words:", len(rows))
    print("saved:", out_path)


if __name__ == "__main__":
    main()
