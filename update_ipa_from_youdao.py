# -*- coding: utf-8 -*-
"""批量用有道词典音标覆盖 vocab / community / Excel。"""
from __future__ import annotations

import json
import shutil
from pathlib import Path

from youdao_ipa import fetch_youdao_ipa_cached

ROOT = Path(__file__).resolve().parent
DOCS = ROOT / "docs"
VOCAB_JSON = DOCS / "vocab-data.json"
COMMUNITY_JSON = DOCS / "community-words.json"
REPO_XLSX = ROOT / "面试常用单词_Java_Python_FastAPI_LangChain.xlsx"
DOCS_XLSX = DOCS / "interview-vocab.xlsx"


def update_vocab_json(cache: dict[str, str]) -> tuple[int, int]:
    raw = json.loads(VOCAB_JSON.read_text(encoding="utf-8"))
    rows = raw.get("rows") or []
    changed = 0
    for row in rows:
        if not row or len(row) < 5:
            continue
        word = str(row[1] or "").strip()
        if not word:
            continue
        ipa = fetch_youdao_ipa_cached(word, cache)
        if not ipa:
            continue
        old_ipa = str(row[4] or "").strip()
        old_pron = str(row[5] or "").strip() if len(row) > 5 else ""
        if old_ipa != ipa or old_pron != ipa:
            row[4] = ipa
            if len(row) > 5:
                row[5] = ipa
            changed += 1
    VOCAB_JSON.write_text(
        json.dumps(raw, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    return changed, len(rows)


def update_community_json(cache: dict[str, str]) -> int:
    if not COMMUNITY_JSON.exists():
        return 0
    raw = json.loads(COMMUNITY_JSON.read_text(encoding="utf-8"))
    words = raw.get("words") if isinstance(raw, dict) else raw
    if not isinstance(words, list):
        return 0
    changed = 0
    for w in words:
        if not isinstance(w, dict):
            continue
        word = str(w.get("word") or "").strip()
        if not word:
            continue
        ipa = fetch_youdao_ipa_cached(word, cache)
        if not ipa:
            continue
        if str(w.get("ipa") or "").strip() != ipa or str(w.get("pron") or "").strip() != ipa:
            w["ipa"] = ipa
            w["pron"] = ipa
            changed += 1
    if isinstance(raw, dict):
        raw["words"] = words
        COMMUNITY_JSON.write_text(json.dumps(raw, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    else:
        COMMUNITY_JSON.write_text(json.dumps(words, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return changed


def update_excel(cache: dict[str, str], path: Path) -> int:
    if not path.exists():
        return 0
    import openpyxl

    wb = openpyxl.load_workbook(path)
    changed = 0
    for name in wb.sheetnames:
        ws = wb[name]
        # 假定第 2 列单词、第 5 列音标、第 6 列发音
        for i in range(2, ws.max_row + 1):
            word = ws.cell(i, 2).value
            if word is None or not str(word).strip():
                continue
            # 跳过表头行误伤
            if str(word).strip() in {"单词", "word", "Word"}:
                continue
            ipa = fetch_youdao_ipa_cached(str(word).strip(), cache)
            if not ipa:
                continue
            old_ipa = ws.cell(i, 5).value
            old_pron = ws.cell(i, 6).value
            if str(old_ipa or "").strip() != ipa or str(old_pron or "").strip() != ipa:
                ws.cell(i, 5).value = ipa
                ws.cell(i, 6).value = ipa
                changed += 1
    wb.save(path)
    wb.close()
    return changed


def main() -> None:
    cache: dict[str, str] = {}
    print("updating vocab-data.json ...")
    v_changed, v_total = update_vocab_json(cache)
    print(f"  vocab: {v_changed}/{v_total} rows updated, cache={len(cache)}")

    print("updating community-words.json ...")
    c_changed = update_community_json(cache)
    print(f"  community: {c_changed} entries updated")

    for path in (REPO_XLSX, DOCS_XLSX):
        print(f"updating {path.name} ...")
        try:
            n = update_excel(cache, path)
            print(f"  excel cells changed: {n}")
        except PermissionError:
            print(f"  locked, skip {path}")
        except Exception as e:
            print(f"  failed {path}: {e}")

    if REPO_XLSX.exists() and DOCS_XLSX.exists():
        try:
            shutil.copy2(REPO_XLSX, DOCS_XLSX)
            print("copied repo xlsx -> docs/interview-vocab.xlsx")
        except PermissionError:
            print("copy skipped (locked)")

    print("done. unique words queried:", len(cache))


if __name__ == "__main__":
    main()
