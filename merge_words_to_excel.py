# -*- coding: utf-8 -*-
"""合并社区投稿到 Excel，并重建 docs/vocab-data.json。"""
from __future__ import annotations

import json
import re
import shutil
import urllib.request
from pathlib import Path

ROOT = Path(__file__).resolve().parent
DESKTOP_XLSX = Path(r"e:\Users\i\Desktop\单词\面试常用单词_Java_Python_FastAPI_LangChain.xlsx")
REPO_XLSX = ROOT / "面试常用单词_Java_Python_FastAPI_LangChain.xlsx"
COMMUNITY_JSON = ROOT / "docs" / "community-words.json"
PENDING_JSON = ROOT / "docs" / "pending-words.json"
DOCS = ROOT / "docs"
ISSUE_REPO = "Tewelve-creator/twelve"

HEADER = [
    "分类", "单词", "用法/解释", "中文谐音", "音标",
    "发音", "技术栈", "重要程度", "对比词", "一句话回答",
]
COMMUNITY_SHEET = "社区添加单词"


def pick(body: str, key: str) -> str:
    m = re.search(rf"[-*]\s*\*\*{re.escape(key)}\*\*:\s*(.+)", body or "")
    return (m.group(1).strip().replace("—", "") if m else "")


def load_community_words() -> list[dict]:
    words: list[dict] = []
    for path in (COMMUNITY_JSON, PENDING_JSON):
        if not path.exists():
            continue
        raw = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(raw, list):
            words.extend(raw)
        elif isinstance(raw, dict):
            words.extend(raw.get("words") or [])

    try:
        url = f"https://api.github.com/repos/{ISSUE_REPO}/issues?labels=word-submission&state=open&per_page=100"
        req = urllib.request.Request(
            url,
            headers={"Accept": "application/vnd.github+json", "User-Agent": "twelve-merge"},
        )
        with urllib.request.urlopen(req, timeout=20) as resp:
            issues = json.loads(resp.read().decode("utf-8"))
        for issue in issues or []:
            if issue.get("pull_request"):
                continue
            body = issue.get("body") or ""
            word = pick(body, "单词") or str(issue.get("title") or "").replace("[词条投稿]", "").strip()
            if not word:
                continue
            trouble = [t for t in pick(body, "卡在哪").replace("，", "、").split("、") if t and t != "（未选）"]
            words.append(
                {
                    "word": word,
                    "explain": pick(body, "用法/解释"),
                    "homo": pick(body, "中文谐音"),
                    "ipa": pick(body, "音标"),
                    "pron": pick(body, "发音"),
                    "tech": pick(body, "技术栈") or "用户投稿",
                    "level": pick(body, "重要程度") or "了解",
                    "compare": pick(body, "对比词") or "—",
                    "oneliner": pick(body, "一句话回答"),
                    "trouble": trouble,
                    "cat": pick(body, "分类") or "",
                }
            )
    except Exception as e:
        print("issues fetch skipped:", e)

    dedup = {}
    for w in words:
        if not isinstance(w, dict):
            continue
        key = str(w.get("word") or "").strip().lower()
        if key:
            dedup[key] = w
    return list(dedup.values())


def enrich_ipa(w: dict, cache: dict | None = None) -> str:
    """优先用已有音标；缺失或明显占位时用有道词典补全。"""
    from youdao_ipa import fetch_youdao_ipa_cached, format_youdao_ipa

    word = str(w.get("word") or "").strip()
    ipa = str(w.get("ipa") or "").strip()
    if ipa and ipa not in {"—", "-", "/—/"} and ("/" in ipa or "英" in ipa or "美" in ipa):
        # 已是规范音标；若只有一侧也可保留
        if "英" in ipa or "美" in ipa or ipa.startswith("/"):
            return ipa
    if not word:
        return ipa
    store = cache if cache is not None else {}
    yd = fetch_youdao_ipa_cached(word, store)
    return yd or ipa or format_youdao_ipa("", "", "")


def word_to_row(w: dict, ipa_cache: dict | None = None) -> list[str]:
    trouble = w.get("trouble") or []
    cat = str(w.get("cat") or "").strip()
    if not cat:
        if isinstance(trouble, list) and trouble:
            cat = "社区·" + "/".join(str(t) for t in trouble if t)
        else:
            cat = "未分类"
    tech = str(w.get("tech") or "用户投稿").strip() or "用户投稿"
    ipa = enrich_ipa(w, ipa_cache)
    pron = str(w.get("pron") or "").strip() or ipa
    return [
        cat,
        str(w.get("word") or "").strip(),
        str(w.get("explain") or "").strip(),
        str(w.get("homo") or "").strip(),
        ipa,
        pron,
        tech,
        str(w.get("level") or "了解").strip() or "了解",
        str(w.get("compare") or w.get("note") or "—").strip() or "—",
        str(w.get("oneliner") or "").strip(),
    ]


def ensure_header_row(ws) -> None:
    if ws.max_row < 1 or all(c.value is None for c in ws[1]):
        ws.append(HEADER)
        return
    # 对齐表头列名
    for i, name in enumerate(HEADER, start=1):
        if ws.cell(1, i).value != name:
            ws.cell(1, i).value = name


def sync_community_sheet(wb, community: list[dict], ipa_cache: dict | None = None) -> int:
    """新建/重建「社区添加单词」表，只存放网站社区投稿。"""
    if COMMUNITY_SHEET in wb.sheetnames:
        ws = wb[COMMUNITY_SHEET]
        wb.remove(ws)
    # 插到「全部汇总」后面，方便查找
    idx = 1
    if "全部汇总" in wb.sheetnames:
        idx = wb.sheetnames.index("全部汇总") + 1
    ws = wb.create_sheet(COMMUNITY_SHEET, idx)
    ws.append(HEADER)
    rows = []
    seen = set()
    cache = ipa_cache if ipa_cache is not None else {}
    for w in community:
        row = word_to_row(w, cache)
        key = row[1].strip().lower()
        if not key or key in seen:
            continue
        seen.add(key)
        rows.append(row)
    # 按单词排序，便于浏览
    rows.sort(key=lambda r: r[1].lower())
    for row in rows:
        ws.append(row)
    return len(rows)


def community_sheet_explains(wb) -> dict[str, str]:
    """word_lower -> explain currently stored on the community sheet."""
    if COMMUNITY_SHEET not in wb.sheetnames:
        return {}
    ws = wb[COMMUNITY_SHEET]
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        if not row or row[1] is None or not str(row[1]).strip():
            continue
        out[str(row[1]).strip().lower()] = str(row[2] or "").strip()
    return out


def looks_like_community_summary_row(cat: str, explain: str, community_explain: str) -> bool:
    cat = (cat or "").strip()
    explain = (explain or "").strip()
    community_explain = (community_explain or "").strip()
    if cat.startswith("社区"):
        return True
    if community_explain and explain == community_explain:
        return True
    return "面试中需能解释其含义" in explain or "技术面试常见术语" in explain


def remove_dropped_community_from_summary(ws, dropped: dict[str, str]) -> int:
    if not dropped:
        return 0
    removed = 0
    for i in range(ws.max_row, 1, -1):
        word = ws.cell(i, 2).value
        if word is None or not str(word).strip():
            continue
        key = str(word).strip().lower()
        if key not in dropped:
            continue
        cat = str(ws.cell(i, 1).value or "")
        explain = str(ws.cell(i, 3).value or "")
        if looks_like_community_summary_row(cat, explain, dropped[key]):
            ws.delete_rows(i)
            removed += 1
    return removed


def merge_into_workbook(xlsx_path: Path, community: list[dict]) -> tuple[int, int, int]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path)
    if "全部汇总" not in wb.sheetnames:
        raise SystemExit(f"missing sheet 全部汇总 in {xlsx_path}")
    ws = wb["全部汇总"]
    ensure_header_row(ws)

    old_community = community_sheet_explains(wb)
    new_keys = {
        str(w.get("word") or "").strip().lower()
        for w in community
        if str(w.get("word") or "").strip()
    }
    dropped = {k: v for k, v in old_community.items() if k not in new_keys}
    removed = remove_dropped_community_from_summary(ws, dropped)
    if removed:
        print(f"removed {removed} deleted community word(s) from 全部汇总")

    existing = set()
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row and len(row) > 1 and row[1] is not None and str(row[1]).strip():
            existing.add(str(row[1]).strip().lower())

    ipa_cache: dict = {}
    added = 0
    for w in community:
        key = str(w.get("word") or "").strip().lower()
        if not key or key in existing:
            continue
        ws.append(word_to_row(w, ipa_cache))
        existing.add(key)
        added += 1

    community_n = sync_community_sheet(wb, community, ipa_cache)

    # 回写 community-words.json 的有道音标，方便站点直接展示
    try:
        for w in community:
            if not isinstance(w, dict):
                continue
            ipa = enrich_ipa(w, ipa_cache)
            if ipa:
                w["ipa"] = ipa
                w["pron"] = ipa
        COMMUNITY_JSON.write_text(
            json.dumps({"updatedAt": __import__("datetime").datetime.utcnow().isoformat() + "Z", "words": community}, ensure_ascii=False, indent=2)
            + "\n",
            encoding="utf-8",
        )
    except Exception as e:
        print("community-words ipa rewrite skipped:", e)

    total = sum(
        1
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True)
        if row and len(row) > 1 and row[1] is not None and str(row[1]).strip()
    )
    wb.save(xlsx_path)
    wb.close()
    return added, total, community_n


def export_vocab_json(xlsx_path: Path) -> int:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["全部汇总"]
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()
    header = ["" if c is None else str(c) for c in (rows_raw[0] if rows_raw else HEADER)][:10]
    while len(header) < 10:
        header.append(HEADER[len(header)])
    rows = []
    for r in rows_raw[1:]:
        if not r or r[1] is None or str(r[1]).strip() == "":
            continue
        rows.append([("" if (i >= len(r) or r[i] is None) else str(r[i]).strip()) for i in range(10)])

    DOCS.mkdir(exist_ok=True)
    (DOCS / "vocab-data.json").write_text(
        json.dumps({"header": header, "rows": rows}, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    shutil.copy2(xlsx_path, DOCS / "interview-vocab.xlsx")
    return len(rows)


def main() -> None:
    community = load_community_words()
    print("community words:", len(community))
    candidates = [p for p in (REPO_XLSX, DESKTOP_XLSX) if p.exists()]
    if not candidates:
        raise SystemExit("No Excel file found")

    primary = None
    last_err = None
    added = total = community_n = 0
    for path in candidates:
        try:
            added, total, community_n = merge_into_workbook(path, community)
            primary = path
            print(f"merged into {path}: +{added} to 全部汇总, total {total}; 社区添加单词 {community_n}")
            break
        except PermissionError as e:
            last_err = e
            print("locked, skip", path)
        except Exception as e:
            last_err = e
            print("failed", path, e)
    if primary is None:
        raise SystemExit(f"cannot write Excel: {last_err}")

    for other in candidates:
        if other.resolve() == primary.resolve():
            continue
        try:
            shutil.copy2(primary, other)
            print("copied to", other)
        except PermissionError:
            print("copy skipped (locked):", other)

    if primary.resolve() != REPO_XLSX.resolve():
        try:
            shutil.copy2(primary, REPO_XLSX)
        except PermissionError:
            pass

    n = export_vocab_json(REPO_XLSX if REPO_XLSX.exists() else primary)
    print("docs/vocab-data.json rows:", n)


if __name__ == "__main__":
    main()
