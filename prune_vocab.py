# -*- coding: utf-8 -*-
"""Prune interview vocab to practical interview-explanation terms.

Keep: design concepts, frequently spoken interview APIs, core stack terms.
Drop: obscure lifecycle rote lists, ultra-basic everyday APIs, rarely spoken jargon.
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

XLSX = Path(__file__).resolve().parent / "面试常用单词_Java_Python_FastAPI_LangChain.xlsx"
ALSO = [
    Path(__file__).resolve().parent / "vocab.xlsx",
    Path(__file__).resolve().parent / "netlify-deploy" / "interview-vocab.xlsx",
]

HEADER = [
    "分类",
    "单词",
    "用法/解释",
    "中文谐音",
    "音标",
    "发音",
    "技术栈",
    "重要程度",
    "对比词",
    "一句话回答",
]

# Always drop (safe globally — not ambiguous with important types like Set/Map/List)
DROP_ALWAYS = {
    # React class lifecycle / obscure
    "componentdidmount",
    "componentdidupdate",
    "componentwillunmount",
    "shouldcomponentupdate",
    "getderivedstatefromprops",
    "getsnapshotbeforeupdate",
    "usedebugvalue",
    "uselayouteffect",
    "render props",
    "higher-order component",
    "storybook",
    "testing library",
    "jest",
    "lazy",
    "memo",
    "suspense",
    # Vue lifecycle rote
    "beforecreate",
    "created",
    "beforemount",
    "mounted",
    "beforeupdate",
    "updated",
    "beforedestroy",
    "destroyed",
    "beforeunmount",
    "unmounted",
    "activated",
    "deactivated",
    "onbeforemount",
    "onmounted",
    "onbeforeupdate",
    "onupdated",
    "onbeforeunmount",
    "onunmounted",
    "onactivated",
    "ondeactivated",
    "onerrorcaptured",
    "scopedslots",
    "listeners",
    "toref",
    "torefs",
    "isref",
    "isreactive",
    "isreadonly",
    "isproxy",
    "definecomponent",
    "transitiongroup",
    "routerview",
    "routerlink",
    "watcheffect",
    "eventbus",
    "scoped slot",
    "render function",
    "devtools",
    "keepalive",
    "transition",
    "nexttick",
    "provide/inject",
    "v-bind",
    "v-on",
    "v-show",
    "v-slot",
    "ssg",
    "nuxt",
    # HTML ultra-basic
    "element",
    "tag",
    "attribute",
    "value",
    "doctype",
    # JS everyday APIs / noise
    "push",
    "pop",
    "shift",
    "unshift",
    "slice",
    "splice",
    "concat",
    "join",
    "reverse",
    "sort",
    "indexof",
    "lastindexof",
    "foreach",
    "find",
    "findindex",
    "some",
    "every",
    "includes",
    "flat",
    "flatmap",
    "getelementbyid",
    "getelementsbyclassname",
    "getelementsbytagname",
    "queryselector",
    "queryselectorall",
    "createelement",
    "appendchild",
    "removechild",
    "replacechild",
    "innerhtml",
    "textcontent",
    "setattribute",
    "getattribute",
    "removeattribute",
    "classlist",
    "addeventlistener",
    "removeeventlistener",
    "preventdefault",
    "stoppropagation",
    "regexp",
    "date",
    "math",
    "parseint",
    "parsefloat",
    "isnan",
    "isfinite",
    "encodeuri",
    "decodeuri",
    "encodeuricomponent",
    "decodeuricomponent",
    "isarray",
    "tostring",
    "tofixed",
    "charat",
    "substring",
    "trim",
    "touppercase",
    "tolowercase",
    "match",
    "replace",
    "split",
    "keys",
    "values",
    "entries",
    "assign",
    "defineproperty",
    "hasownproperty",
    "instanceof",
    "typeof",
    "debugger",
    "void",
    "yield",
    "extends",
    "arrow",
    "xmlhttprequest",
    "canvas",
    "webgl",
    "memoize",
    "curry",
    "angular",
    "babel",
    "eslint",
    "prettier",
    "undefined",
    "arrow function",
    "bind",
    "call",
    "apply",
    "symbol",
    "reflect",
    "memoization",
    "bubble sort",
    "observer",
    "publisher",
    "subscriber",
    "webpack",
    # Python ultra-basic
    "str",
    "int",
    "float",
    "bool",
    "none",
    "bytes",
    "variable",
    "script",
    "self",
    "init",
    "open",
    "with",
    "raise",
    "try-except",
    "black",
    "virtualenv",
    "interpreter",
    "cpython",
    "pep",
    "indentation",
    "comprehension",
    # FastAPI low-value
    "redoc",
    "htmlresponse",
    "startup/shutdown",
    "sync def",
    "summary/description",
    "tags",
    "include router",
    "mount",
    "staticfiles",
    "jsonresponse",
    "request validation error",
    "exception handler",
    "annotated",
    # LangChain / AI tools rarely useful as vocab drills
    "messagesplaceholder",
    "model factory",
    "chat model",
    "langsmith",
    "opencode",
    "sandbox",
    "repo indexing",
    "diff apply",
    "terminal agent",
    "cursor vs copilot",
    "code review with ai",
    "pair programming with ai",
    "hallucinated api",
    # OpenCV rarely spoken
    "haar cascade",
    "sift",
    "orb",
    "warpaffine",
    "perspective transform",
    "morphology",
    "imshow",
    "crop",
    # Java low-value
    "vo",
    "nacos",
    "rabbitmq",
    "jre",
    "gradle",
    "junit",
    "servlet",
    "runnable",
    "throws",
    "try-catch",
    "transient",
    "hashset",
    "linkedlist",
    "serializable",
}

# Drop only in frontend-ish categories / 网页原文
DROP_FRONTEND_ONLY = {
    "el",
    "options",
    "parent",
    "root",
    "children",
    "data",
    "slots",
    "refs",
    "emit",
    "watch",
    "readonly",
    "mixin",
    "plugin",
    "cli",
    "instance",
    "directive",
    "getter",
    "mutation",
    "action",
    "navigation",
    "route",
    "router",
    "slot",
    "style",
    "try",
    "catch",
    "finally",
    "throw",
    "new",
    "delete",
    "super",
    "static",
    "private",
    "public",
    "get",
    "set",
    "then",
    "import",
    "export",
    "default",
    "require",
    "module",
    "let",
    "const",
    "template",
    "type",
    "enum",
    "function",
    "array",
    "object",
    "string",
    "number",
    "boolean",
    "null",
    "constructor",
    "event",
    "listener",
    "this",
    "iterator",
    "proxy",
    "search",
    "map",  # array map in JS sheet
    "filter",
    "reduce",
}

# Drop FastAPI generic param names when dedicated "path parameter" etc. exist
DROP_FASTAPI_GENERIC = {
    "path",
    "query",
    "body",
    "header",
    "form",
    "file",
    "template",
}

# Drop Java ultra-basic when OOP/collection sheets already cover concepts
DROP_JAVA_BASIC = {
    "package",
    "import",
    "protected",
    "final",
    "collection",
    "annotation",
}

FORCE_KEEP = {
    "react",
    "vue",
    "dom",
    "html",
    "virtual dom",
    "props",
    "state",
    "usestate",
    "useeffect",
    "usecontext",
    "usereducer",
    "usecallback",
    "usememo",
    "useref",
    "context",
    "ref",
    "redux",
    "next.js",
    "code splitting",
    "tree shaking",
    "debouncing",
    "throttling",
    "debounce",
    "throttle",
    "promise",
    "async",
    "await",
    "prototype",
    "json",
    "api",
    "ajax",
    "cookie",
    "localstorage",
    "sessionstorage",
    "websocket",
    "fetch",
    "typescript",
    "node.js",
    "spa",
    "ssr",
    "vite",
    "pinia",
    "vuex",
    "composition api",
    "reactive",
    "computed",
    "setup",
    "v-model",
    "v-if",
    "v-for",
    "guard",
    "vue router",
    "react router",
    "error boundary",
    "component",
    "jsx",
    "semantic",
    "accessibility",
    "seo",
    "bom",
    "callback",
    "scope",
    "class",
    "stack",
    "tree",
    "hash",
    "quick sort",
    "merge sort",
    "generator",
    "closure",
    "event loop",
    "rest",
    "jwt",
    "cors",
    "middleware",
    "https",
    "oauth2",
    "spring",
    "spring boot",
    "jvm",
    "jdk",
    "java",
    "python",
    "fastapi",
    "langchain",
    "langgraph",
    "rag",
    "llm",
    "embedding",
    "token",
    "prompt",
    "transformer",
    "opencv",
    "cursor",
    "redis",
    "kafka",
    "nginx",
    "deadlock",
    "thread",
    "exception",
    "encapsulation",
    "inheritance",
    "polymorphism",
    "abstraction",
    "interface",
    "override",
    "overload",
    "volatile",
    "synchronized",
    "arraylist",
    "hashmap",
    "concurrenthashmap",
    "decorator",
    "gil",
    "pydantic",
    "depends",
    "agent",
    "tool",
    "checkpoint",
    "hallucination",
    "context window",
    "temperature",
    "list",
    "dict",
    "tuple",
    "stream",
    "lambda",
    "generics",
    "reflection",
    "mybatis",
    "maven",
    "ioc",
    "di",
    "aop",
    "gc",
    "heap",
    "transaction",
    "mvcc",
    "dto",
    "entity",
    "path parameter",
    "query parameter",
    "request body",
}


def norm(word: str) -> str:
    return re.sub(r"\s+", " ", (word or "").strip().lower())


def is_frontend_context(sheet: str, category: str) -> bool:
    if sheet == "网页原文":
        return True
    return category in {"React", "Vue", "JavaScript常用单词", "未分类"}


def should_drop(
    word: str,
    category: str,
    sheet: str,
    importance: str,
    usage: str = "",
) -> bool:
    w = norm(word)
    if not w:
        return True
    usage_l = (usage or "").lower()

    if is_frontend_context(sheet, category):
        if w in {"map", "set"} and ("集合" in (usage or "") or "键值" in (usage or "")):
            return False
        if w in {"map", "filter", "reduce"} and category == "JavaScript常用单词":
            if "映射数组" in (usage or "") or "过滤数组" in (usage or "") or "缩减" in (usage or ""):
                return True
        if w in DROP_FRONTEND_ONLY:
            # allow data-structure Map/Set already handled
            if w in {"map", "set"} and ("集合" in (usage or "") or "键值" in (usage or "")):
                return False
            return True

    if w in FORCE_KEEP:
        return False

    if w in DROP_ALWAYS:
        return True

    if sheet in {"FastAPI常用单词", "FastAPI补充"} and w in DROP_FASTAPI_GENERIC:
        return True

    if sheet in {"Java常用单词", "Java后端补充"} and w in DROP_JAVA_BASIC:
        return True

    if re.match(
        r"^(component|get|should)(did|will|before|derived|snapshot)",
        w.replace(" ", ""),
    ):
        return True

    return False


def dedupe_rows(rows: list[list]) -> list[list]:
    seen: set[str] = set()
    out = []
    for r in rows:
        key = norm(str(r[1]))
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def style_header(ws):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for col, name in enumerate(HEADER, 1):
        cell = ws.cell(1, col, name)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws):
    widths = [12, 28, 42, 22, 22, 18, 14, 10, 22, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def filter_sheet_rows(rows: list[list], sheet: str) -> tuple[list[list], list[str]]:
    kept, removed = [], []
    for r in rows:
        while len(r) < 10:
            r.append("")
        r = ["" if c is None else c for c in r]
        cat, word, usage, imp = str(r[0]), str(r[1]), str(r[2]), str(r[7])
        if should_drop(word, cat, sheet, imp, usage):
            removed.append(f"{sheet}\t{cat}\t{word}\t{imp}")
        else:
            kept.append(r[:10])
    return dedupe_rows(kept), removed


def rewrite_sheet(wb, name: str, rows: list[list], index: int | None = None):
    if name in wb.sheetnames:
        idx = wb.sheetnames.index(name)
        wb.remove(wb[name])
    else:
        idx = index if index is not None else len(wb.sheetnames)
    nws = wb.create_sheet(name, idx)
    style_header(nws)
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            cell = nws.cell(i, j, val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    autosize(nws)


def main():
    wb = load_workbook(XLSX)
    # snapshot original order
    original_order = list(wb.sheetnames)

    all_removed: list[str] = []
    per_sheet_stats: dict[str, tuple[int, int]] = {}
    filtered: dict[str, list[list]] = {}

    content_sheets = [s for s in original_order if s not in {"全部汇总", "说明"}]
    for name in content_sheets:
        ws = wb[name]
        rows = [[c if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]
        body = [list(r) for r in rows[1:]] if rows else []
        kept, removed = filter_sheet_rows(body, name)
        all_removed.extend(removed)
        per_sheet_stats[name] = (len(body), len(kept))
        filtered[name] = kept

    # Rebuild content sheets in original relative order
    for name in content_sheets:
        rewrite_sheet(wb, name, filtered[name])

    # Rebuild 全部汇总
    merged: list[list] = []
    seen: set[str] = set()
    order = [
        "网页原文",
        "Java常用单词",
        "Python常用单词",
        "FastAPI常用单词",
        "LangChain常用单词",
        "Python补充",
        "FastAPI补充",
        "LangChain与RAG补充",
        "Transformer补充",
        "OpenCV补充",
        "AI编程工具补充",
        "Java后端补充",
    ]
    for name in order:
        for row in filtered.get(name, []):
            key = norm(str(row[1]))
            if key in seen:
                continue
            seen.add(key)
            merged.append(row)

    rewrite_sheet(wb, "全部汇总", merged, index=0)
    per_sheet_stats["全部汇总"] = (0, len(merged))

    # Ensure 说明 stays last if present
    if "说明" in wb.sheetnames:
        ws = wb["说明"]
        wb.move_sheet(ws, offset=len(wb.sheetnames) - 1 - wb.sheetnames.index("说明"))

    wb.save(XLSX)
    for p in ALSO:
        if p.exists() and p.resolve() != XLSX.resolve():
            try:
                wb.save(p)
            except Exception as e:
                print(f"skip save {p}: {e}")

    report = Path(__file__).resolve().parent / "_prune_report.txt"
    lines = ["=== Prune stats ==="]
    for name in ["全部汇总"] + content_sheets:
        before, after = per_sheet_stats[name]
        if name == "全部汇总":
            lines.append(f"{name}: kept={after}")
        else:
            lines.append(f"{name}: {before} -> {after} (removed {before - after})")
    lines.append(f"\nTotal removed row occurrences: {len(all_removed)}")
    lines.append("\n=== All removed ===")
    lines.extend(all_removed)
    report.write_text("\n".join(lines), encoding="utf-8")
    print("\n".join(lines[:35]))
    print(f"... removed total {len(all_removed)}")
    print(f"Report: {report}")
    print(f"Saved: {XLSX}")


if __name__ == "__main__":
    main()
