# -*- coding: utf-8 -*-
"""把《AI与Java后端面试词汇分类表》合并进全部汇总，并重建 vocab-data.json。"""
from __future__ import annotations

import json
import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
SRC = Path(r"e:\Users\i\Desktop\AI与Java后端面试词汇分类表.xlsx")
REPO_XLSX = ROOT / "面试常用单词_Java_Python_FastAPI_LangChain.xlsx"
DESKTOP_XLSX = Path(r"e:\Users\i\Desktop\单词\面试常用单词_Java_Python_FastAPI_LangChain.xlsx")
DOCS = ROOT / "docs"
HEADER = [
    "分类", "单词", "用法/解释", "中文谐音", "音标",
    "发音", "技术栈", "重要程度", "对比词", "一句话回答",
]
LEVEL_MAP = {"核心": "必问", "高频": "常问", "了解": "了解"}


def cells(row) -> list[str]:
    vals = ["" if c is None else str(c).strip() for c in row[:10]]
    while len(vals) < 10:
        vals.append("")
    return vals


def load_classified() -> dict[str, list[str]]:
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb["全部词汇"]
    out = {}
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = cells(raw)
        word = row[1]
        if not word:
            continue
        row[7] = LEVEL_MAP.get(row[7], row[7] or "了解")
        out[word.lower()] = row
    wb.close()
    return out


def merge_sheet(xlsx_path: Path, classified: dict[str, list[str]]) -> tuple[int, int]:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["全部汇总"]
    for i, name in enumerate(HEADER, start=1):
        ws.cell(1, i).value = name

    existing = {}
    for idx, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = cells(raw)
        key = row[1].lower()
        if key:
            existing[key] = idx

    updated = added = 0
    for key, row in classified.items():
        if key in existing:
            r = existing[key]
            for c, val in enumerate(row, start=1):
                ws.cell(r, c).value = val
            updated += 1
        else:
            ws.append(row)
            added += 1
    wb.save(xlsx_path)
    wb.close()
    return added, updated


def export_json(xlsx_path: Path) -> int:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["全部汇总"]
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()
    rows = []
    for r in rows_raw[1:]:
        if not r or r[1] is None or str(r[1]).strip() == "":
            continue
        rows.append([("" if (i >= len(r) or r[i] is None) else str(r[i]).strip()) for i in range(10)])
    DOCS.mkdir(exist_ok=True)
    (DOCS / "vocab-data.json").write_text(
        json.dumps({"header": HEADER, "rows": rows}, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    shutil.copy2(xlsx_path, DOCS / "interview-vocab.xlsx")
    shutil.copy2(SRC, ROOT / SRC.name)
    shutil.copy2(SRC, DOCS / SRC.name)
    return len(rows)


def main() -> None:
    classified = load_classified()
    print("classified:", len(classified))
    added, updated = merge_sheet(REPO_XLSX, classified)
    print(f"repo excel +{added} updated {updated}")
    try:
        shutil.copy2(REPO_XLSX, DESKTOP_XLSX)
        print("copied desktop excel")
    except PermissionError:
        print("desktop excel locked, skip copy")
    n = export_json(REPO_XLSX)
    print("vocab-data.json rows:", n)


if __name__ == "__main__":
    main()
